# ⚠️ import streamlit as st 가 완전히 삭제된 순수 파이썬/FastAPI용 백엔드 모듈입니다.
import requests
import urllib.parse
import json
import xml.etree.ElementTree as ET
import openpyxl
import html 
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from dateutil.relativedelta import relativedelta

seoul_gu_pool = {
    '11110': '종로구', '11140': '중구', '11170': '용산구', '11200': '성동구', '11215': '광진구',
    '11230': '동대문구', '11260': '중랑구', '11290': '성북구', '11305': '강북구', '11320': '도봉구',
    '11350': '노원구', '11380': '은평구', '11410': '서대문구', '11440': '마포구', '11470': '양천구',
    '11500': '강서구', '11530': '구로구', '11545': '금천구', '11560': '영등포구', '11590': '동작구',
    '11620': '관악구', '11650': '서초구', '11680': '강남구', '11710': '송파구', '11740': '강동구'
}

seoul_dong_pool = {
    '11710': ['잠실동', '신천동', '풍납동', '송파동', '석촌동', '삼전동', '가락동', '문정동', '장지동', '방이동', '오금동', '거여동', '마천동'],
    '11530': ['신도림동', '구로동', '가리봉동', '고척동', '개봉동', '오류동', '궁동', '온수동', '천왕동', '항동']
}

def clean_apt_name(name):
    if not name: return ""
    return name.replace(" ", "").replace("(주상복합)", "").replace("(도시형생활주택)", "").replace("주상복합", "").replace("아파트", "")

def extract_items(response_text):
    text = response_text.strip()
    if not text: return []
    if text.startswith('{'):
        try:
            data = json.loads(text)
            body = data.get('response', {}).get('body', {})
            if not isinstance(body, dict): return []
            if 'item' in body:
                items_node = body['item']
            else:
                items_parent = body.get('items', {})
                items_node = items_parent.get('item', []) if isinstance(items_parent, dict) else items_parent
            if isinstance(items_node, dict):
                return [items_node]
            elif isinstance(items_node, list):
                return items_node
            else:
                return []
        except:
            return []
    elif text.startswith('<'):
        try:
            root = ET.fromstring(text)
            return root.findall('.//item')
        except:
            return []
    return []

def get_field(item, *keys):
    if isinstance(item, dict):
        for k in keys:
            if k in item and item[k] is not None: return str(item[k]).strip()
        return ""
    else:
        for k in keys:
            val = item.findtext(k)
            if val is not None: return val.strip()
        return ""

def generate_excel_data(api_key, district_code, district_name, target_dong, start_date, end_date, apt_filters):
    log_text = ""
    yield "progress", "🔄 데이터 추출을 시작합니다..."

    api_key = urllib.parse.unquote(api_key.strip())
    url_apt_list_sigungu = 'https://apis.data.go.kr/1613000/AptListService3/getSigunguAptList3'
    url_apt_list_bjd = 'https://apis.data.go.kr/1613000/AptListService3/getLegaldongAptList3'
    url_apt_info = 'https://apis.data.go.kr/1613000/AptBasisInfoServiceV4/getAphusBassInfoV4'

    trade_api_urls = [
        'https://apis.data.go.kr/1613000/RTMSDataSvcAptTrade/getRTMSDataSvcAptTrade',
        'https://apis.data.go.kr/1613000/RTMSDataSvcAptPreSaleRightTrade/getRTMSDataSvcAptPreSaleRightTrade'
    ]

    kapt_name_to_code = {}
    apt_details_map = {}
    deals_filtered_map = {}
    unique_bjd_codes = set()

    count_trade_api = 0
    count_kapt_api = 0

    master_scan_months = []
    current = datetime(start_date.year, start_date.month, 1)
    end_month_calc = datetime(end_date.year, end_date.month, 1)

    while current <= end_month_calc:
        master_scan_months.append(current.strftime("%Y%m"))
        current += relativedelta(months=1)

    safe_target_dong = html.escape(str(target_dong or ''))
    dong_log = f"'{safe_target_dong}' 전체" if target_dong != "전체" else "구 전체"
    log_text += f"\n🔄 [1단계] {master_scan_months[0]} ~ {master_scan_months[-1]} 기간 {dong_log} 실거래가 추출 중...\n"
    yield "log", log_text

    for deal_ymd in master_scan_months:
        for url_endpoint in trade_api_urls:
            trade_params = {'serviceKey': api_key, 'LAWD_CD': district_code, 'DEAL_YMD': deal_ymd, 'numOfRows': '10000'}
            try:
                res_trade = requests.get(url_endpoint, params=trade_params, timeout=10)
                count_trade_api += 1

                if "LIMITED NUMBER OF SERVICE REQUESTS EXCEEDS ERROR" in res_trade.text:
                    yield "error", "국토교통부 실거래가 API 일일 호출 한도를 초과했습니다."
                    return

                if res_trade.status_code == 200:
                    items = extract_items(res_trade.text)
                    for item in items:
                        try:
                            apt_nm = get_field(item, 'aptNm', '아파트명')
                            exclu_ar_str = get_field(item, 'excluUseAr', '전용면적')
                            dong = get_field(item, 'umdNm', '법정동')

                            if not apt_nm or not exclu_ar_str: continue
                            if target_dong != "전체" and target_dong not in dong: continue

                            if apt_filters:
                                clean_trade_apt = apt_nm.replace(" ", "")
                                matched = any(f.replace(" ", "") in clean_trade_apt for f in apt_filters)
                                if not matched: continue

                            s_code = get_field(item, 'sggCd', 'sigunguCd')
                            d_code = get_field(item, 'umdCd', 'eubmyundongCd')
                            if len(s_code) == 5 and len(d_code) == 5:
                                unique_bjd_codes.add(s_code + d_code)
                            elif len(d_code) == 5:
                                unique_bjd_codes.add(district_code + d_code)

                            exclu_ar = float(exclu_ar_str)
                            d_year = get_field(item, 'dealYear', '년')
                            d_month = get_field(item, 'dealMonth', '월').zfill(2)
                            d_day = get_field(item, 'dealDay', '일').zfill(2)

                            safe_district_name = html.escape(str(district_name or ''))
                            safe_dong = html.escape(str(dong or ''))
                            region_key = f"{safe_district_name} ({safe_dong})"
                            group_key = (region_key, apt_nm, exclu_ar)

                            deal_date_full = f"{d_year}.{d_month}.{d_day}"
                            amt_val = int(get_field(item, 'dealAmount', '거래금액').replace(',', '')) * 10

                            if group_key not in deals_filtered_map: deals_filtered_map[group_key] = []
                            deals_filtered_map[group_key].append({'deal_amount': amt_val, 'deal_date': deal_date_full})
                        except Exception:
                            continue
            except Exception:
                continue

    if not deals_filtered_map:
        yield "error", "설정하신 조건(기간/동/필터)에 해당하는 실거래 내역이 없습니다."
        return

    log_text += f"\n🏢 [2단계] K-APT 인덱싱 스캔 시작...\n"
    yield "log", log_text

    for bjd_code in unique_bjd_codes:
        list_params_bjd = {'serviceKey': api_key, 'bjdCode': bjd_code, 'numOfRows': '9999', 'pageNo': '1'}
        try:
            res_list = requests.get(url_apt_list_bjd, params=list_params_bjd, timeout=5)
            count_kapt_api += 1
            for item in extract_items(res_list.text):
                kcode = get_field(item, 'kaptCode')
                if kcode: kapt_name_to_code[clean_apt_name(get_field(item, 'kaptName'))] = kcode
        except:
            pass

    list_params_sig = {'serviceKey': api_key, 'sigunguCode': district_code, 'numOfRows': '9999', 'pageNo': '1'}
    try:
        res_list = requests.get(url_apt_list_sigungu, params=list_params_sig, timeout=10)
        count_kapt_api += 1
        for item in extract_items(res_list.text):
            kcode = get_field(item, 'kaptCode')
            if kcode: kapt_name_to_code[clean_apt_name(get_field(item, 'kaptName'))] = kcode
    except:
        pass

    for apt_nm in set([key[1] for key in deals_filtered_map.keys()]):
        trade_cleaned_key = clean_apt_name(apt_nm)
        target_kcode = kapt_name_to_code.get(trade_cleaned_key)

        if not target_kcode:
            for k_clean, k_code in kapt_name_to_code.items():
                if trade_cleaned_key in k_clean or k_clean in trade_cleaned_key:
                    target_kcode = k_code
                    break

        if target_kcode:
            try:
                info_params = {'serviceKey': api_key, 'kaptCode': target_kcode}
                res_info = requests.get(url_apt_info, params=info_params, timeout=5)
                count_kapt_api += 1

                if res_info.status_code == 200:
                    items = extract_items(res_info.text)
                    if items:
                        info_item = items[0]
                        raw_h_cnt = get_field(info_item, 'kaptdaCnt')
                        if not raw_h_cnt or raw_h_cnt in ['0', '0.0']:
                            alt_cnt = get_field(info_item, 'hoCnt')
                            if alt_cnt: raw_h_cnt = alt_cnt

                        if raw_h_cnt and raw_h_cnt != "-":
                            try:
                                raw_h_cnt = str(int(float(raw_h_cnt)))
                            except:
                                raw_h_cnt = "-"
                        else:
                            raw_h_cnt = "-"

                        use_dt = get_field(info_item, 'kaptUsedate')
                        if use_dt and len(use_dt) >= 8:
                            parsed_dt = f"{use_dt[:4]}.{use_dt[4:6]}.{use_dt[6:8]}"
                        else:
                            parsed_dt = "-"

                        cls_type = get_field(info_item, 'codeAptNm')
                        classification = "주상복합" if "주상복합" in cls_type else ""

                        apt_details_map[trade_cleaned_key] = {'households': raw_h_cnt, 'move_in': parsed_dt,
                                                              'classification': classification}
                        
                        safe_apt_nm = html.escape(str(apt_nm or ''))
                        safe_parsed_dt = html.escape(str(parsed_dt or ''))
                        safe_raw_h_cnt = html.escape(str(raw_h_cnt or ''))
                        log_text += f"✅ K-APT 매칭 완료: {safe_apt_nm} (입주:{safe_parsed_dt} | 세대:{safe_raw_h_cnt})\n"
                        yield "log", log_text
            except:
                pass
        else:
            safe_apt_nm = html.escape(str(apt_nm or ''))
            log_text += f"❌ K-APT 매칭 실패: {safe_apt_nm}\n"
            yield "log", log_text

        if trade_cleaned_key not in apt_details_map:
            apt_details_map[trade_cleaned_key] = {'households': "-", 'move_in': "-", 'classification': ""}

    log_text += f"\n📊 [3단계] 엑셀 파일 생성 중...\n"
    yield "log", log_text

    sorted_keys = sorted(deals_filtered_map.keys(), key=lambda x: (x[0], x[1], x[2]))
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "실거래 상세 분석"

    font_header = Font(name='맑은 고딕', size=10, bold=True)
    font_body = Font(name='맑은 고딕', size=10)
    font_summary = Font(name='맑은 고딕', size=10, bold=True, color='000000')
    fill_header = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    fill_summary = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_right = Alignment(horizontal='right', vertical='center')
    border_thin = Side(border_style="thin", color="D3D3D3")
    cell_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    headers = ["구분", "단지명", "입주시기", "세대수", "공급면적(㎡)", "전용면적(㎡)", "공급(py)", "전용(py)", "전용률", "매매가", "공급평단가", "전용평단가", "계약일자"]
    ws.append(headers)

    for col_idx in range(1, 14):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = cell_border

    current_row = 2

    for key in sorted_keys:
        region_text, apt_nm, exclu_ar = key
        trade_cleaned_key = clean_apt_name(apt_nm)
        info_p = apt_details_map.get(trade_cleaned_key, {})

        move_in_text = info_p.get('move_in', "-")
        household_text = info_p.get('households', "-")
        cls_type = info_p.get('classification', "")

        display_apt_name = apt_nm
        if cls_type == "주상복합" and "주상복합" not in display_apt_name:
            display_apt_name += "(주상복합)"
        elif cls_type == "도시형생활주택" and "도시형" not in display_apt_name:
            display_apt_name += "(도시형생활주택)"

        deals = deals_filtered_map[key]
        deals.sort(key=lambda x: x['deal_date'], reverse=True)

        for deal in deals:
            final_deal_amount = deal['deal_amount']
            final_deal_date = deal['deal_date']

            supply_m2 = round(exclu_ar / 0.76, 2)
            supply_py = round(supply_m2 * 0.3025, 2)
            exclu_py = round(exclu_ar * 0.3025, 2)
            ratio = (exclu_ar / supply_m2) if supply_m2 > 0 else 0

            supply_price = int(round(final_deal_amount / supply_py, 0)) if supply_py > 0 else 0
            exclu_price = int(round(final_deal_amount / exclu_py, 0)) if exclu_py > 0 else 0

            h_val = household_text
            if h_val.isdigit(): h_val = int(h_val)

            row_data = [region_text, display_apt_name, str(move_in_text), h_val, supply_m2, exclu_ar, supply_py,
                        exclu_py, ratio, final_deal_amount, supply_price, exclu_price, final_deal_date]
            ws.append(row_data)

            for col_idx in range(1, 14):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.font = font_body
                cell.border = cell_border
                cell.alignment = align_center if col_idx in [1, 2, 3, 4, 13] else align_right
                if col_idx in [5, 6, 7, 8]:
                    cell.number_format = '#,##0.00'
                elif col_idx == 9:
                    cell.number_format = '0%'
                elif col_idx == 4 and isinstance(h_val, int):
                    cell.number_format = '#,##0'
                elif col_idx in [10, 11, 12]:
                    cell.number_format = '#,##0'

            current_row += 1

        amounts = [d['deal_amount'] for d in deals]
        avg_amt = sum(amounts) / len(amounts)

        supply_m2 = round(exclu_ar / 0.76, 2)
        supply_py = round(supply_m2 * 0.3025, 2)
        exclu_py = round(exclu_ar * 0.3025, 2)

        avg_supply_price = int(round(avg_amt / supply_py, 0)) if supply_py > 0 else 0
        avg_exclu_price = int(round(avg_amt / exclu_py, 0)) if exclu_py > 0 else 0

        summary_row = [
            region_text, display_apt_name, f"▶ {exclu_ar}㎡ 요약", None, None, None, None, None, None,
            f"최소: {min(amounts):,}\n최대: {max(amounts):,}\n평균: {int(avg_amt):,}",
            f"평균: {avg_supply_price:,}", f"평균: {avg_exclu_price:,}", None
        ]
        ws.append(summary_row)
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row, end_column=9)

        for col_idx in range(1, 14):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.border = cell_border
            if col_idx >= 3:
                cell.fill = fill_summary
                cell.font = font_summary
            else:
                cell.font = font_body
            if col_idx <= 3 or col_idx >= 10:
                cell.alignment = align_center
        current_row += 1

    merge_ranges = []
    start_r_A = 2
    for r in range(3, current_row + 1):
        if (ws.cell(row=r, column=1).value if r < current_row else None) != ws.cell(row=start_r_A, column=1).value:
            if r - 1 > start_r_A: merge_ranges.append((start_r_A, 1, r - 1, 1))
            start_r_A = r

    start_r_B = 2
    for r in range(3, current_row + 1):
        if (ws.cell(row=r, column=2).value if r < current_row else None) != ws.cell(row=start_r_B, column=2).value or \
                (ws.cell(row=r, column=1).value if r < current_row else None) != ws.cell(row=start_r_B, column=1).value:
            if r - 1 > start_r_B: merge_ranges.append((start_r_B, 2, r - 1, 2))
            start_r_B = r

    for m in merge_ranges:
        for r_idx in range(m[0] + 1, m[2] + 1): ws.cell(row=r_idx, column=m[1]).value = None
        ws.merge_cells(start_row=m[0], start_column=m[1], end_row=m[2], end_column=m[3])
        ws.cell(row=m[0], column=m[1]).alignment = align_center

    for col in ws.columns:
        ws.column_dimensions[get_column_letter(col[0].column)].width = 16

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['J'].width = 25
    ws.column_dimensions['K'].width = 18
    ws.column_dimensions['L'].width = 18

    output = BytesIO()
    wb.save(output)

    log_text += f"\n🎉 엑셀 대시보드 생성이 완료되었습니다!\n🧾 API 사용량: 실거래가 {count_trade_api}회 / K-APT {count_kapt_api}회"
    yield "log", log_text

    yield "success", {
        "data": output.getvalue(),
        "filename": f"[{district_name}_{target_dong}]_실거래_분석.xlsx"
    }
